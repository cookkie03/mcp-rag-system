"""Estrattori di testo per formati non-testuali (PDF, Audio, Notebook, Excel)"""

import logging
from pathlib import Path

logger = logging.getLogger(__name__)


def extract_pdf_text(file_path: Path, mode: str = "markdown") -> str:
    """
    Estrae testo da PDF con struttura preservata.
    
    Args:
        file_path: Path del PDF
        mode: "layout" (analisi avanzata), "markdown" (tabelle) o "text" (legacy)
        
    Returns:
        Testo estratto (Markdown strutturato)
        
    Modes:
        - "layout": Analisi avanzata con pymupdf_layout (blocchi, colonne, tabelle)
        - "markdown": pymupdf4llm (tabelle, headers, liste)
        - "text": PyMuPDF base (solo testo lineare)
    """
    
    # === MODE: LAYOUT (analisi avanzata) ===
    if mode == "layout":
        try:
            import fitz
            from pymupdf_layout import LayoutAnalyzer
            
            doc = fitz.open(file_path)
            analyzer = LayoutAnalyzer()
            
            all_text = []
            for page_num, page in enumerate(doc, 1):
                # Analizza il layout della pagina
                layout = analyzer.analyze(page)
                
                page_parts = []
                for block in layout:
                    block_type = getattr(block, 'type', 'paragraph')
                    text = getattr(block, 'text', str(block)).strip()
                    
                    if not text:
                        continue
                    
                    # Formatta in base al tipo di blocco
                    if block_type == 'header':
                        page_parts.append(f"## {text}")
                    elif block_type == 'table':
                        page_parts.append(f"\n{text}\n")
                    else:
                        page_parts.append(text)
                
                if page_parts:
                    all_text.append(f"<!-- Page {page_num} -->\n" + "\n\n".join(page_parts))
            
            doc.close()
            return "\n\n---\n\n".join(all_text).strip()
            
        except ImportError as e:
            logger.warning(f"pymupdf_layout import error: {e}")
            logger.warning("pymupdf_layout non installato o fallito, fallback a pymupdf4llm")
            mode = "markdown"  # Fallback
        except Exception as e:
            logger.warning(f"pymupdf_layout fallito ({e}), fallback a pymupdf4llm")
            mode = "markdown"  # Fallback
    
    # === MODE: MARKDOWN (pymupdf4llm) ===
    if mode == "markdown":
        try:
            import pymupdf4llm
            # Estrae come Markdown con tabelle, headers, liste
            md_text = pymupdf4llm.to_markdown(str(file_path))
            return md_text.strip() if md_text else ""
        except ImportError:
            logger.warning("pymupdf4llm non installato, fallback a PyMuPDF base")
        except Exception as e:
            logger.warning(f"pymupdf4llm fallito ({e}), fallback a PyMuPDF base")
    
    # === FALLBACK: estrazione base con PyMuPDF ===
    import fitz
    text_parts = []
    with fitz.open(file_path) as doc:
        for page in doc:
            text_parts.append(page.get_text())
    return "\n\n".join(text_parts).strip()


def extract_audio_text(file_path: Path, model_name: str = "base") -> str:
    """Trascrive audio usando Whisper (richiede FFmpeg)"""
    import whisper
    
    logger.info(f"Caricamento modello Whisper '{model_name}'...")
    model = whisper.load_model(model_name)
    
    logger.info(f"Trascrizione {file_path.name}...")
    result = model.transcribe(str(file_path))
    
    return result["text"].strip()


def extract_notebook_text(file_path: Path) -> str:
    """Estrae testo da Jupyter Notebook (code + markdown cells)"""
    import json
    nb = json.loads(file_path.read_text(encoding='utf-8'))
    texts = []
    for cell in nb.get('cells', []):
        source = cell.get('source', [])
        texts.append(''.join(source) if isinstance(source, list) else source)
    return '\n\n'.join(texts).strip()


def extract_excel_text(file_path: Path) -> str:
    """Estrae testo da Excel (.xlsx)"""
    from openpyxl import load_workbook
    wb = load_workbook(file_path, read_only=True, data_only=True)
    texts = []
    for sheet in wb:
        for row in sheet.iter_rows(values_only=True):
            texts.append(' '.join(str(c) for c in row if c is not None))
    return '\n'.join(texts).strip()


def extract_text(file_path: Path, config: dict = None) -> tuple[str, str | None]:
    """Estrae testo da qualsiasi formato supportato.
    
    Args:
        file_path: Path del file da processare
        config: Configurazione (opzionale, per whisper_model)
        
    Returns:
        (testo, errore) - se errore è None, l'estrazione è riuscita
    """
    ext = file_path.suffix.lower()
    
    # Ottieni estensioni da config o usa default
    if config:
        pdf_ext = set(config.get('extensions', {}).get('pdf', ['.pdf']))
        audio_ext = set(config.get('extensions', {}).get('audio', ['.mp3', '.m4a', '.wav', '.ogg', '.flac']))
        notebook_ext = set(config.get('extensions', {}).get('notebook', ['.ipynb']))
        excel_ext = set(config.get('extensions', {}).get('excel', ['.xlsx']))
        whisper_model = config.get('whisper_model', 'base')
        pdf_mode = config.get('pdf_extraction_mode', 'markdown')
    else:
        pdf_ext = {'.pdf'}
        audio_ext = {'.mp3', '.m4a', '.wav', '.ogg', '.flac'}
        notebook_ext = {'.ipynb'}
        excel_ext = {'.xlsx'}
        whisper_model = 'base'
        pdf_mode = 'markdown'
    
    try:
        if ext in pdf_ext:
            return extract_pdf_text(file_path, mode=pdf_mode), None
        elif ext in audio_ext:
            return extract_audio_text(file_path, whisper_model), None
        elif ext in notebook_ext:
            return extract_notebook_text(file_path), None
        elif ext in excel_ext:
            return extract_excel_text(file_path), None
        else:
            # File di testo normale
            return file_path.read_text(encoding='utf-8', errors='ignore').strip(), None
    except ImportError as e:
        return "", f"Dipendenza mancante: {e}"
    except Exception as e:
        logger.error(f"Errore estrazione {file_path}: {e}")
        return "", str(e)
